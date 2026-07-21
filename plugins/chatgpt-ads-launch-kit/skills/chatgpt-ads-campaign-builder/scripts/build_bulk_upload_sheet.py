#!/usr/bin/env python3
"""
Build a ChatGPT Ads BULK UPLOAD workbook (.xlsx) that matches the platform's
own bulk upload template field for field: three tabs (campaigns, adgroups,
ads), exact header casing, no filler or instruction rows. This file is meant
to be uploaded to ChatGPT Ads as-is through the bulk upload feature, not
reformatted first.

Usage:
    python build_bulk_upload_sheet.py campaign.json "Brand ChatGPT Ads Bulk Upload.xlsx"

Requires openpyxl:
    pip install openpyxl   (add --break-system-packages if your environment needs it)

Input JSON schema
------------------
{
  "campaign": {
    "campaign_name": "brandtag0620",          # alphanumeric only, used as an ID
    "budget_max": 17,
    "budget_type": "Daily",                    # "Lifetime" or "Daily", Daily is the safer default
    "launch_date": "2026-06-20",
    "end_date": "",                             # optional, leave "" for an ongoing test
    "objective": "Views",                       # "Views" or "Clicks"
    "target_countries": ["US"]                  # subset of AU, CA, GB, JP, KR, NZ, US
  },
  "adgroups": [
    {
      "adgroup_name": "brandtag0620gen",       # alphanumeric only, used as an ID
      "max_bid": 12,                            # CPM if objective is Views, CPC if Clicks
      "context_hints": ["frozen fruit snack", "healthy dessert", "..."],
      "ads": [
        {
          "title": "...",       # max 24 chars
          "copy": "...",        # max 48 chars
          "link": "https://example.com/?utm_source=openai&utm_medium=cpc&...",
          "image_link": "https://drive.google.com/uc?export=view&id=..."
        }
      ]
    }
  ]
}

Notes
-----
- campaign_name and adgroup_name are literal foreign keys across tabs in the
  real template. Keep them alphanumeric only, no dashes, spaces, or
  punctuation, or the platform's importer will not be able to link the tabs.
- target_countries is only allowed to contain AU, CA, GB, JP, KR, NZ, US as
  of this template version.
- objective "Clicks" is listed by the platform, but this specific template
  file (named a CPM template) notes clicks campaigns were rolling out in
  beta at the time it was issued. Confirm Clicks is live for the account in
  the ChatGPT Ads dashboard before shipping a CPC campaign.
- image_link must be a live, publicly viewable URL (square image, 640 to
  1200px, PNG or JPG), hosted anywhere available (Google Drive with link
  sharing turned on, Imgur, Dropbox, a site's own media library). Do not
  assume Drive or any other specific storage tool is connected to Claude.
  This script only checks that the link isn't empty or still a placeholder,
  it can't confirm the link actually resolves publicly.
"""

import json
import re
import sys

import openpyxl

ALLOWED_COUNTRIES = {"AU", "CA", "GB", "JP", "KR", "NZ", "US"}
ID_PATTERN = re.compile(r"^[A-Za-z0-9]+$")


def validate(spec):
    errors = []

    c = spec["campaign"]
    if not ID_PATTERN.match(c["campaign_name"]):
        errors.append(
            f"campaign_name '{c['campaign_name']}' must be alphanumeric only, no spaces or punctuation."
        )
    if c["objective"] not in ("Views", "Clicks"):
        errors.append("objective must be 'Views' or 'Clicks'.")
    bad_countries = [x for x in c["target_countries"] if x not in ALLOWED_COUNTRIES]
    if bad_countries:
        errors.append(
            f"target_countries {bad_countries} not supported. Allowed: {sorted(ALLOWED_COUNTRIES)}."
        )

    adgroup_names = set()
    for ag in spec["adgroups"]:
        if not ID_PATTERN.match(ag["adgroup_name"]):
            errors.append(
                f"adgroup_name '{ag['adgroup_name']}' must be alphanumeric only, no spaces or punctuation."
            )
        if ag["adgroup_name"] in adgroup_names:
            errors.append(f"Duplicate adgroup_name '{ag['adgroup_name']}'.")
        adgroup_names.add(ag["adgroup_name"])

        if not ag.get("ads"):
            errors.append(f"Ad group '{ag['adgroup_name']}' has no ads.")

        for ad in ag.get("ads", []):
            if len(ad["title"]) > 24:
                errors.append(f"Ad title '{ad['title']}' is {len(ad['title'])} chars, max is 24.")
            if len(ad["copy"]) > 48:
                errors.append(f"Ad copy '{ad['copy']}' is {len(ad['copy'])} chars, max is 48.")
            image_link = ad.get("image_link", "")
            if not image_link or "REPLACE_WITH_FILE_ID" in image_link or "PLACEHOLDER" in image_link.upper():
                errors.append(
                    f"Ad '{ad['title']}' has a missing or placeholder image_link. "
                    f"Host the creative somewhere public and use its real link before this file is upload-ready."
                )

    return errors


def build(spec, output_path):
    errors = validate(spec)

    wb = openpyxl.Workbook()
    c = spec["campaign"]

    # --- campaigns tab ---
    ws = wb.active
    ws.title = "campaigns"
    ws.append(
        ["campaign_name", "budget_max", "budget_type", "launch_date", "end_date", "objective", "target_countries"]
    )
    ws.append(
        [
            c["campaign_name"],
            c["budget_max"],
            c["budget_type"],
            c["launch_date"],
            c.get("end_date") or "",
            c["objective"],
            json.dumps(c["target_countries"], separators=(",", ":")),
        ]
    )

    # --- adgroups tab ---
    ws2 = wb.create_sheet("adgroups")
    ws2.append(["campaign_name", "adgroup_name", "max_bid", "Context Hints"])
    for ag in spec["adgroups"]:
        ws2.append(
            [
                c["campaign_name"],
                ag["adgroup_name"],
                ag["max_bid"],
                json.dumps(ag["context_hints"], separators=(",", ":")),
            ]
        )

    # --- ads tab ---
    ws3 = wb.create_sheet("ads")
    ws3.append(["adgroup_name", "title", "copy", "link", "image_link"])
    for ag in spec["adgroups"]:
        for ad in ag.get("ads", []):
            ws3.append(
                [
                    ag["adgroup_name"],
                    ad["title"],
                    ad["copy"],
                    ad["link"],
                    ad["image_link"],
                ]
            )

    wb.save(output_path)
    return errors


def main():
    if len(sys.argv) != 3:
        print("Usage: python build_bulk_upload_sheet.py campaign.json output.xlsx")
        sys.exit(1)

    with open(sys.argv[1]) as f:
        spec = json.load(f)

    errors = build(spec, sys.argv[2])
    print(f"Wrote {sys.argv[2]}")
    if errors:
        print("\nFix these before uploading to ChatGPT Ads:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
