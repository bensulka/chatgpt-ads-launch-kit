#!/usr/bin/env python3
"""
Build a ChatGPT Ads campaign launch workbook (.xlsx) with three tabs:
Campaign, Ad Copy, and Context Hints. The workbook is self-contained and
opens in Excel, Google Sheets, or Numbers.

Usage:
    python build_campaign_sheet.py campaign.json output.xlsx

Requires openpyxl:
    pip install openpyxl   (add --break-system-packages if your environment needs it)

Input JSON schema
-----------------
{
  "brand": "Example Brand",
  "phase": "Phase 1",
  "campaign": {
    "name": "June Launch Test - Clicks - 6/20/26",
    "budget": "$500",
    "budget_type": "Daily - about $17/day",
    "objective": "Clicks",
    "destination_url": "https://example.com"
  },
  "ad_groups": [
    {
      "name": "General - Home Page - V1",
      "destination_urls": [
        "https://example.com?utm_source=openai&utm_medium=cpc&utm_campaign=ChatGPTAdsTest&utm_content=GeneralV-A",
        "https://example.com?utm_source=openai&utm_medium=cpc&utm_campaign=ChatGPTAdsTest&utm_content=GeneralV-B"
      ],
      "versions": [
        {"label": "Version A (Long)",  "headline": "...", "description": "..."},
        {"label": "Version B (Short)", "headline": "...", "description": "..."}
      ],
      "context_hints": ["frozen fruit snack", "healthy dessert", "..."]
    }
  ]
}

Notes
-----
- destination_urls may have 1 entry (single URL) or 2 (A/B with different
  utm_content). The script labels the row "Destination URL" for one and
  "Destination URL(s)" for two.
- Headline and Description character counts are computed automatically.
"""

import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Light, clean styling
TITLE_FONT = Font(bold=True, size=12)
SECTION_FONT = Font(bold=True, size=11)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # header row fill
LABEL_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def build_campaign_tab(ws, data):
    brand = data.get("brand", "")
    phase = data.get("phase", "Phase 1")
    c = data["campaign"]

    ws["A1"] = f"{brand} | ChatGPT Ads | {phase}"
    ws["A1"].font = TITLE_FONT

    rows = [
        ("Campaign Name", c.get("name", "")),
        ("Budget", c.get("budget", "")),
        ("Budget Type", c.get("budget_type", "")),
        ("Objective", c.get("objective", "")),
        ("Destination URL", c.get("destination_url", "")),
    ]
    r = 2
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value)
        r += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60


def build_ad_copy_tab(ws, data):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12

    r = 1
    for group in data["ad_groups"]:
        # Section header = ad group name
        cell = ws.cell(row=r, column=1, value=group["name"])
        cell.font = SECTION_FONT
        r += 1

        ws.cell(row=r, column=1, value="Ad Group Name").font = LABEL_FONT
        ws.cell(row=r, column=2, value=group["name"])
        r += 1

        urls = group.get("destination_urls", [])
        if len(urls) <= 1:
            ws.cell(row=r, column=1, value="Destination URL").font = LABEL_FONT
            ws.cell(row=r, column=2, value=urls[0] if urls else "")
            r += 1
        else:
            ws.cell(row=r, column=1, value="Destination URL(s)").font = LABEL_FONT
            ws.cell(row=r, column=2, value=urls[0])
            r += 1
            for u in urls[1:]:
                ws.cell(row=r, column=2, value=u)
                r += 1

        # Copy table header
        headers = ["Version", "Headline", "Description", "Headline Chars", "Desc Chars"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        r += 1

        for v in group.get("versions", []):
            headline = v.get("headline", "")
            desc = v.get("description", "")
            ws.cell(row=r, column=1, value=v.get("label", ""))
            ws.cell(row=r, column=2, value=headline).alignment = WRAP
            ws.cell(row=r, column=3, value=desc).alignment = WRAP
            ws.cell(row=r, column=4, value=len(headline))
            ws.cell(row=r, column=5, value=len(desc))
            r += 1

        r += 1  # blank spacer row between ad groups


def build_context_hints_tab(ws, data):
    groups = data["ad_groups"]
    for i, group in enumerate(groups):
        col = i * 2 + 1  # 1, 3, 5 ... separator column between each
        name_cell = ws.cell(row=1, column=col, value=group["name"])
        name_cell.font = SECTION_FONT
        hdr = ws.cell(row=2, column=col, value="Context Hint")
        hdr.font = HEADER_FONT
        hdr.fill = HEADER_FILL
        for j, hint in enumerate(group.get("context_hints", []), start=3):
            ws.cell(row=j, column=col, value=hint).alignment = TOP
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 48
        if i < len(groups) - 1:
            sep = ws.cell(row=1, column=col + 1).column_letter
            ws.column_dimensions[sep].width = 4


def main():
    if len(sys.argv) != 3:
        print("Usage: python build_campaign_sheet.py campaign.json output.xlsx")
        sys.exit(1)

    with open(sys.argv[1]) as f:
        data = json.load(f)

    wb = Workbook()
    campaign_ws = wb.active
    campaign_ws.title = "Campaign"
    build_campaign_tab(campaign_ws, data)

    build_ad_copy_tab(wb.create_sheet("Ad Copy"), data)
    build_context_hints_tab(wb.create_sheet("Context Hints"), data)

    wb.save(sys.argv[2])
    print(f"Wrote {sys.argv[2]}")


if __name__ == "__main__":
    main()
